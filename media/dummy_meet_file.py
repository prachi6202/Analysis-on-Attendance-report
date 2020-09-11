import numpy as np
import openpyxl as op
import xlwt
from xlwt import Workbook
file = op.load_workbook("C:\\Users\prachi\Desktop\meet_attedence.xlsx")

att_sheet=file["Sheet1"]
name_sheet=file["Sheet2"]

row=att_sheet.max_row
col=att_sheet.max_column

att_list=[]
for i in range(3,row+1):
    att_list.append(att_sheet.cell(row=i,column=2).value)


att_list = list(filter(None,att_list))
att_list = np.unique(att_list)
att_list.sort()


dt=att_sheet.cell(row=1,column=2).value


stu_len = name_sheet.max_row

student_list = []
for i in range(2,stu_len+1):
    student_list.append(name_sheet.cell(row=i,column=2).value)

student_list = list(filter(None,student_list))
student_list = np.unique(student_list)
student_list.sort()

print(len(att_list),"  ",len(student_list))

for i in range(0,len(att_list)-1):
    #print(i,"   ",att_list[i])
    pass

print(" ************************************** ")

for i in range(0,len(student_list)-1):
    #print(i,"   ",student_list[i])
    pass



attendance=[dt.date(),]

for i in range(len(student_list)):
    x = student_list[i]
    if(x in att_list):
        attendance.append(1)
        
    else:
        attendance.append(0)
       
print(len(student_list),len(attendance)-1,len(att_list))

for i in range(1,len(attendance)):
    print(attendance[i])


##wb=Workbook()
##result=wb.add_sheet('Sheet2')
##for i in range(len(attendance)):
##    result.write(i,3,attendance[i])
##
##wb.save("C:\\Users\prachi\Desktop\meet_attendence.xls")

print(attendance)

print(len(attendance))

