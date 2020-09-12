import numpy as np
import openpyxl as op
file = op.load_workbook("C:\\Users\prachi\Desktop\meet_attedence.xlsx")
#print(file.sheet1)

sheet=file["Sheet1"]
# var=sheet.cell(row=1,column=2).value
#print(var)

#print(file['A2'].value)

#print(sheet.max_row)
#print(sheet.max_column)

i=1
row=sheet.max_row
col=sheet.max_column
list1=[]
list2=[]
for i in range(3,row+1):
    list2.append(sheet.cell(row=i,column=2).value)

list2 = list(filter(None,list2))
list2=np.unique(list2)
print(list2)
print(len(list2))
##today=datetime.datetime.now()
##print(today.date())

while(i<col):
    dt=sheet.cell(row=1,column=i).value
    #print("datetime  ",dt.date(),"  type  ",type(dt))
    meetid=sheet.cell(row=2,column=i).value

    
    
    for j in range(3,row+1):
        list1.append((sheet.cell(row=j,column=i).value))
        

    #print("loop ",i," completed")
    
    i+=1
    dt_next=sheet.cell(row=1,column=i).value
    meetid_next=sheet.cell(row=2,column=i).value

    #print(dt.date(),dt_next.date())
    #print(meetid , meetid_next)
    
    if(dt.date()==dt_next.date() and meetid == meetid_next):
        pass
    else:
        break
        #meetid=meetid_next

list1 = list(filter(None,list1))
list1=np.unique(list1)
print(meetid)
for i in range(len(list1)):
    #print(list1[i])
    pass

#print(len(list1))




